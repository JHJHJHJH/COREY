#!/usr/bin/env python3
"""Extract the CORNET X COP and industry workbook into deterministic JSONL artifacts."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import fitz
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_PDF = ROOT / "docs/official-cx/corenet-x-cop---3-1-edition-2025-12.pdf"
DEFAULT_XLSX = ROOT / "docs/official-cx/industry-mapping-4-dec-2025139335b79c8943d695c7b84984c9d50b.xlsx"
DEFAULT_OUT = ROOT / ".cache/knowledge/corenet-x-3.1"
COP_OFFICIAL_URL = "https://info.corenet.gov.sg/regulatory-process/corenet-x-code-of-practice"
MAPPING_OFFICIAL_URL = "https://info.corenet.gov.sg/ifc-sg/requirements---submission/ifc-sg-excel-mapping-file"
SPACE_VALUE_GROUPS = {
    "OccupancyType",
    "SpaceName",
    "AGF_DevelopmentUse",
    "AGF_BonusGFAType",
    "AGF_Name",
    "AGF_BuildingTypology",
    "ALS_LandscapeType",
    "ALS_GreeneryFeatures",
    "AGF_SupportingFacility",
    "AST_AreaType",
}
PLACEHOLDERS = {
    "",
    "n.a",
    "n.a.",
    "na",
    "nil",
    "none",
    "please refer to property sets below",
    "all subtypes listed in cop",
}
DISCIPLINE_FILL_COLORS = {
    (0.929, 0.49, 0.192): "Architecture",
    (0.439, 0.678, 0.278): "C&S",
    (1.0, 0.753, 0.0): "M&E",
}


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def stable_id(*parts: Any) -> str:
    body = "\x1f".join(str(part) for part in parts)
    return hashlib.sha256(body.encode("utf-8")).hexdigest()[:32]


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def canonical(value: Any) -> str:
    return normalize(value).casefold()


def json_value(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def write_jsonl(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


class CorpusBuilder:
    def __init__(self, revision_id: str, embedding_model: str) -> None:
        self.revision_id = revision_id
        self.embedding_model = embedding_model
        self.documents: list[dict[str, Any]] = []
        self.evidence: list[dict[str, Any]] = []
        self.chunks: list[dict[str, Any]] = []
        self.nodes: dict[str, dict[str, Any]] = {}
        self.edges: dict[str, dict[str, Any]] = {}
        self.warnings: list[dict[str, Any]] = []

    def add_document(
        self,
        source_key: str,
        source_kind: str,
        title: str,
        file_name: str,
        sha256: str,
        edition: str | None,
        metadata: dict[str, Any],
    ) -> str:
        document_id = stable_id(self.revision_id, "document", source_key)
        self.documents.append(
            {
                "id": document_id,
                "revisionId": self.revision_id,
                "sourceKey": source_key,
                "sourceKind": source_kind,
                "title": title,
                "fileName": file_name,
                "sha256": sha256,
                "edition": edition,
                "metadata": metadata,
            }
        )
        self.add_node("Document", source_key, title, {"documentId": document_id})
        return document_id

    def add_evidence(
        self,
        document_id: str,
        ordinal: int,
        section_path: list[str],
        locator: dict[str, Any],
        raw_text: str,
        metadata: dict[str, Any] | None = None,
    ) -> str | None:
        normalized_text = normalize(raw_text)
        if not normalized_text:
            return None
        evidence_id = stable_id(
            self.revision_id,
            "evidence",
            document_id,
            ordinal,
            json.dumps(locator, sort_keys=True),
            normalized_text,
        )
        self.evidence.append(
            {
                "id": evidence_id,
                "revisionId": self.revision_id,
                "documentId": document_id,
                "ordinal": ordinal,
                "sectionPath": section_path,
                "locator": locator,
                "rawText": raw_text,
                "normalizedText": normalized_text,
                "contentHash": sha256_bytes(normalized_text.encode("utf-8")),
                "metadata": metadata or {},
            }
        )
        return evidence_id

    def add_chunk(
        self,
        document_id: str,
        stable_key: str,
        content: str,
        source_kind: str,
        source_role: str,
        locator: str,
        metadata: dict[str, Any],
        evidence_ids: list[str] | None = None,
        *,
        page_number: int | None = None,
        sheet_name: str | None = None,
        row_start: int | None = None,
        row_end: int | None = None,
    ) -> str | None:
        content = normalize(content)
        if not content:
            return None
        content_hash = sha256_bytes(content.encode("utf-8"))
        chunk_id = stable_id(self.revision_id, "chunk", stable_key, content_hash)
        self.chunks.append(
            {
                "id": chunk_id,
                "revisionId": self.revision_id,
                "documentId": document_id,
                "stableKey": stable_key,
                "contentHash": content_hash,
                "content": content,
                "sourceKind": source_kind,
                "sourceRole": source_role,
                "locator": locator,
                "pageNumber": page_number,
                "sheetName": sheet_name,
                "rowStart": row_start,
                "rowEnd": row_end,
                "tokenCount": max(1, (len(content) + 3) // 4),
                "metadata": metadata,
                "embeddingModel": self.embedding_model,
                "evidenceIds": sorted(set(evidence_ids or [])),
            }
        )
        return chunk_id

    def add_node(self, node_type: str, key: str, label: str, properties: dict[str, Any] | None = None) -> str:
        normalized_key = f"{node_type}:{canonical(key)}"
        node_id = stable_id(self.revision_id, "node", normalized_key)
        existing = self.nodes.get(node_id)
        if existing:
            existing["properties"].update(properties or {})
            return node_id
        self.nodes[node_id] = {
            "id": node_id,
            "revisionId": self.revision_id,
            "canonicalKey": normalized_key,
            "nodeType": node_type,
            "label": normalize(label),
            "properties": properties or {},
        }
        return node_id

    def add_edge(
        self,
        from_node_id: str,
        to_node_id: str,
        relation_type: str,
        assertion_kind: str,
        evidence_chunk_id: str | None,
        properties: dict[str, Any] | None = None,
    ) -> str:
        edge_id = stable_id(
            self.revision_id,
            "edge",
            from_node_id,
            to_node_id,
            relation_type,
            assertion_kind,
        )
        self.edges[edge_id] = {
            "id": edge_id,
            "revisionId": self.revision_id,
            "fromNodeId": from_node_id,
            "toNodeId": to_node_id,
            "relationType": relation_type,
            "evidenceChunkId": evidence_chunk_id,
            "assertionKind": assertion_kind,
            "properties": properties or {},
        }
        return edge_id


def pdf_source_role(page_number: int) -> str:
    if 33 <= page_number <= 208:
        return "requirement"
    if 247 <= page_number <= 248:
        return "glossary"
    if 251 <= page_number <= 439:
        return "ifc_property_guidance"
    if 209 <= page_number <= 250:
        return "preparation_guidance"
    return "source_note"


def page_heading(toc: list[list[Any]], page_number: int) -> str:
    headings = [normalize(item[1]) for item in toc if int(item[2]) <= page_number]
    return headings[-1] if headings else f"Page {page_number}"


def chunk_words(text: str, target_words: int = 650, overlap_words: int = 75) -> list[str]:
    words = text.split()
    if len(words) <= target_words:
        return [text]
    result: list[str] = []
    start = 0
    while start < len(words):
        end = min(len(words), start + target_words)
        result.append(" ".join(words[start:end]))
        if end == len(words):
            break
        start = max(start + 1, end - overlap_words)
    return result


def pdf_block_payload(block: dict[str, Any]) -> tuple[str, float, bool]:
    spans = [span for line in block.get("lines", []) for span in line.get("spans", [])]
    text = " ".join(normalize(span.get("text")) for span in spans).strip()
    largest = max((float(span.get("size", 0)) for span in spans), default=0)
    bold = any("bold" in str(span.get("font", "")).casefold() for span in spans)
    return text, largest, bold


def table_text(rows: list[list[Any]]) -> str:
    rendered = []
    for row in rows:
        cells = [normalize(cell) for cell in row]
        if any(cells):
            rendered.append(" | ".join(cells))
    return "\n".join(rendered)


def requirement_discipline_bands(page: fitz.Page) -> list[tuple[fitz.Rect, str]]:
    bands: list[tuple[fitz.Rect, str]] = []
    for drawing in page.get_drawings():
        fill = drawing.get("fill")
        if not fill:
            continue
        color = tuple(round(component, 3) for component in fill)
        discipline = DISCIPLINE_FILL_COLORS.get(color)
        rect = fitz.Rect(drawing["rect"])
        # The discipline encoding is a narrow vertical band at the left of a
        # requirement row. Excluding the horizontal legend swatches prevents
        # false associations while retaining bands that span several rows.
        if discipline and rect.width <= 20 and rect.height >= 10 and rect.x1 <= 65:
            bands.append((rect, discipline))
    return bands


def split_multi(value: str) -> list[str]:
    if canonical(value) in PLACEHOLDERS:
        return []
    values = re.split(r"\s*(?:\n|;|,|\bor\b)\s*", normalize(value), flags=re.IGNORECASE)
    return [item for item in (normalize(part) for part in values) if item and canonical(item) not in PLACEHOLDERS]


def extract_pdf(builder: CorpusBuilder, path: Path, file_hash: str) -> None:
    pdf = fitz.open(path)
    document_id = builder.add_document(
        "corenet-x-cop-3.1",
        "cop_pdf",
        "CORENET X Code of Practice, Edition 3.1",
        path.name,
        file_hash,
        "3.1 (December 2025)",
        {
            "pageCount": pdf.page_count,
            "tocEntries": len(pdf.get_toc()),
            "path": str(path.relative_to(ROOT)),
            "officialUrl": COP_OFFICIAL_URL,
        },
    )
    document_node = builder.add_node("Document", "corenet-x-cop-3.1", "CORENET X Code of Practice, Edition 3.1")
    toc = pdf.get_toc()
    table_count = 0
    internal_links = 0
    external_links = 0
    evidence_ordinal = 0

    for page_index in range(pdf.page_count):
        page_number = page_index + 1
        page = pdf[page_index]
        heading = page_heading(toc, page_number)
        role = pdf_source_role(page_number)
        section_node = builder.add_node("Section", f"cop:{heading}", heading, {"firstSeenPage": page_number})
        builder.add_edge(document_node, section_node, "CONTAINS", "structure", None)

        page_evidence: list[str] = []
        blocks = sorted(
            page.get_text("dict").get("blocks", []),
            key=lambda block: (block.get("bbox", [0, 0])[1], block.get("bbox", [0, 0])[0]),
        )
        for block_index, block in enumerate(blocks):
            block_text, font_size, bold = pdf_block_payload(block)
            bbox = block.get("bbox")
            if not block_text or not isinstance(bbox, (list, tuple)) or len(bbox) != 4:
                continue
            evidence_ordinal += 1
            evidence_id = builder.add_evidence(
                document_id,
                evidence_ordinal,
                ["CORENET X Code of Practice, Edition 3.1", heading],
                {
                    "page": page_number,
                    "bbox": [round(float(value), 2) for value in bbox],
                },
                block_text,
                {"blockIndex": block_index, "fontSize": round(font_size, 2), "bold": bold},
            )
            if evidence_id:
                page_evidence.append(evidence_id)

        page_text = normalize(page.get_text("text"))
        for part_index, part in enumerate(chunk_words(page_text)):
            builder.add_chunk(
                document_id,
                f"pdf:p{page_number}:text:{part_index}",
                f"{heading}. {part}",
                "cop_pdf",
                role,
                f"COP p. {page_number}",
                {"heading": heading, "part": part_index},
                evidence_ids=page_evidence,
                page_number=page_number,
            )

        for link in page.get_links():
            if link.get("kind") == fitz.LINK_GOTO:
                internal_links += 1
            elif link.get("uri"):
                external_links += 1

        if page_number < 33 or page_number > 439:
            continue
        try:
            tables = page.find_tables().tables
        except Exception as error:  # malformed vector geometry must not abort the corpus
            builder.warnings.append({"source": "pdf", "page": page_number, "message": str(error)})
            continue
        discipline_bands = requirement_discipline_bands(page) if role == "requirement" else []
        for table_index, table in enumerate(tables):
            rows = table.extract()
            content = table_text(rows)
            if not content:
                continue
            table_count += 1
            table_bbox = fitz.Rect(table.bbox)
            disciplines = sorted(
                {
                    discipline
                    for band, discipline in discipline_bands
                    if band.y0 < table_bbox.y1 and band.y1 > table_bbox.y0
                }
            )
            evidence_ordinal += 1
            table_evidence_id = builder.add_evidence(
                document_id,
                evidence_ordinal,
                ["CORENET X Code of Practice, Edition 3.1", heading],
                {
                    "page": page_number,
                    "bbox": [round(float(value), 2) for value in table_bbox],
                },
                content,
                {"tableIndex": table_index, "rowCount": len(rows)},
            )
            chunk_id = builder.add_chunk(
                document_id,
                f"pdf:p{page_number}:table:{table_index}",
                f"{heading}. {content}",
                "cop_pdf",
                role,
                f"COP p. {page_number}, table {table_index + 1}",
                {
                    "heading": heading,
                    "tableIndex": table_index,
                    "rowCount": len(rows),
                    "disciplines": disciplines,
                },
                evidence_ids=[table_evidence_id] if table_evidence_id else [],
                page_number=page_number,
            )
            if role != "ifc_property_guidance" or not rows:
                continue
            headers = [canonical(cell) for cell in rows[0]]
            def column(*needles: str) -> int | None:
                for index, header in enumerate(headers):
                    if any(needle in header for needle in needles):
                        return index
                return None
            entity_col = column("ifc entity")
            subtype_col = column("subtype")
            property_col = column("ifc/sg property", "ifc + sg property", "property name")
            unit_col = column("unit")
            if entity_col is None and property_col is None:
                continue
            component_node = builder.add_node("IdentifiedComponent", heading, heading)
            builder.add_edge(section_node, component_node, "IDENTIFIES_COMPONENT", "cop_guidance", chunk_id)
            for row in rows[1:]:
                values = [normalize(cell) for cell in row]
                entity = values[entity_col] if entity_col is not None and entity_col < len(values) else ""
                prop = values[property_col] if property_col is not None and property_col < len(values) else ""
                unit = values[unit_col] if unit_col is not None and unit_col < len(values) else ""
                if entity:
                    entity_node = builder.add_node("IFCEntity", entity, entity)
                    builder.add_edge(component_node, entity_node, "MAPS_TO_ENTITY", "cop_guidance", chunk_id)
                if subtype_col is not None and subtype_col < len(values):
                    for subtype in split_multi(values[subtype_col]):
                        subtype_node = builder.add_node("IFCSubtype", subtype, subtype)
                        builder.add_edge(component_node, subtype_node, "MAPS_TO_SUBTYPE", "cop_guidance", chunk_id)
                if prop:
                    property_node = builder.add_node("Property", prop, prop)
                    builder.add_edge(component_node, property_node, "REQUIRES_PROPERTY", "cop_guidance", chunk_id)
                    if unit:
                        unit_node = builder.add_node("Unit", unit, unit)
                        builder.add_edge(property_node, unit_node, "USES_UNIT", "cop_guidance", chunk_id)

    builder.documents[-1]["metadata"].update(
        {"tableCount": table_count, "internalLinkCount": internal_links, "externalLinkCount": external_links}
    )
    pdf.close()


def row_content(headers: list[str], values: list[Any]) -> str:
    pairs = []
    for header, value in zip(headers, values):
        rendered = normalize(value)
        if header and rendered:
            pairs.append(f"{header}: {rendered}")
    return ". ".join(pairs)


def row_fields(headers: list[str], values: list[Any]) -> list[dict[str, str]]:
    fields: list[dict[str, str]] = []
    for header, value in zip(headers, values):
        rendered = normalize(value)
        if header and rendered:
            fields.append({"label": header, "value": rendered})
    return fields


def workbook_cell_value(cell: Any) -> Any:
    return json_value(cell.value)


def add_mapping_graph(builder: CorpusBuilder, chunk_id: str | None, row: dict[str, str]) -> None:
    component = row.get("Identified Component", "")
    if not component:
        return
    component_node = builder.add_node("IdentifiedComponent", component, component, {"raw": component})
    agency = row.get("Agency", "")
    if agency:
        agency_node = builder.add_node("Agency", agency, agency)
        builder.add_edge(component_node, agency_node, "REQUESTED_BY", "industry_mapping", chunk_id)
    discipline = row.get("Suggested Discipline", "")
    for value in split_multi(discipline):
        node = builder.add_node("Discipline", value, value)
        builder.add_edge(component_node, node, "REQUIRES_DISCIPLINE", "industry_mapping", chunk_id)
    for value in split_multi(row.get("IFC4 Entities", "")):
        node = builder.add_node("IFCEntity", value, value)
        builder.add_edge(component_node, node, "MAPS_TO_ENTITY", "industry_mapping", chunk_id)
    for value in split_multi(row.get("IFC subtypes (* USERDEFINED)", "")):
        node = builder.add_node("IFCSubtype", value, value)
        builder.add_edge(component_node, node, "MAPS_TO_SUBTYPE", "industry_mapping", chunk_id)
    pset = row.get("Property Set", "")
    prop = row.get("Property Name", "")
    if pset and canonical(pset) not in PLACEHOLDERS:
        node = builder.add_node("PropertySet", pset, pset)
        builder.add_edge(component_node, node, "USES_PROPERTY_SET", "industry_mapping", chunk_id)
    if prop and canonical(prop) not in PLACEHOLDERS:
        property_node = builder.add_node("Property", prop, prop)
        builder.add_edge(component_node, property_node, "REQUIRES_PROPERTY", "industry_mapping", chunk_id)
        unit = row.get("Unit", "")
        if unit and canonical(unit) not in PLACEHOLDERS:
            unit_node = builder.add_node("Unit", unit, unit)
            builder.add_edge(property_node, unit_node, "USES_UNIT", "industry_mapping", chunk_id)
        accepted = row.get("Accepted Values", "")
        if accepted and canonical(accepted) not in PLACEHOLDERS:
            domain_node = builder.add_node("ValueDomain", f"{component}:{prop}", f"{component} · {prop}")
            builder.add_edge(property_node, domain_node, "HAS_ALLOWED_VALUE", "industry_mapping", chunk_id, {"raw": accepted})


def extract_workbook(builder: CorpusBuilder, path: Path, file_hash: str) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)
    document_id = builder.add_document(
        "corenet-x-industry-mapping-2025-12",
        "industry_workbook",
        "CORENET X Industry Mapping",
        path.name,
        file_hash,
        "1 December 2025 alignment to COP 3.1",
        {
            "sheets": workbook.sheetnames,
            "path": str(path.relative_to(ROOT)),
            "officialUrl": MAPPING_OFFICIAL_URL,
        },
    )
    document_node = builder.add_node("Document", "corenet-x-industry-mapping-2025-12", "CORENET X Industry Mapping")
    evidence_ordinal = 0

    def add_row_evidence(
        sheet_name: str,
        row_start: int,
        row_end: int,
        content: str,
        fields: list[dict[str, str]],
    ) -> str | None:
        nonlocal evidence_ordinal
        evidence_ordinal += 1
        return builder.add_evidence(
            document_id,
            evidence_ordinal,
            ["CORENET X Industry Mapping", sheet_name],
            {"sheet": sheet_name, "rowStart": row_start, "rowEnd": row_end},
            content,
            {"structuredFields": fields},
        )

    readme = workbook["README"]
    for row_number in range(1, readme.max_row + 1):
        values = [workbook_cell_value(readme.cell(row_number, column)) for column in range(1, 3)]
        if not any(normalize(value) for value in values):
            continue
        content = " | ".join(normalize(value) for value in values)
        evidence_id = add_row_evidence(
            "README",
            row_number,
            row_number,
            content,
            row_fields(["Topic", "Details"], values),
        )
        chunk_id = builder.add_chunk(
            document_id,
            f"xlsx:README:r{row_number}",
            content,
            "industry_workbook",
            "source_note",
            f"README row {row_number}",
            {},
            evidence_ids=[evidence_id] if evidence_id else [],
            sheet_name="README",
            row_start=row_number,
            row_end=row_number,
        )
        if chunk_id:
            note_node = builder.add_node("Section", f"workbook:readme:{row_number}", f"README row {row_number}")
            builder.add_edge(document_node, note_node, "CONTAINS", "structure", chunk_id)

    change_log = workbook["Change Log"]
    change_headers = [normalize(change_log.cell(1, column).value) for column in range(1, 8)]
    for row_number in range(2, change_log.max_row + 1):
        values = [workbook_cell_value(change_log.cell(row_number, column)) for column in range(1, 8)]
        content = row_content(change_headers, values)
        evidence_id = add_row_evidence(
            "Change Log",
            row_number,
            row_number,
            content,
            row_fields(change_headers, values),
        )
        builder.add_chunk(
            document_id,
            f"xlsx:ChangeLog:r{row_number}",
            content,
            "industry_workbook",
            "change_log",
            f"Change Log row {row_number}",
            {},
            evidence_ids=[evidence_id] if evidence_id else [],
            sheet_name="Change Log",
            row_start=row_number,
            row_end=row_number,
        )

    mapping = workbook["CX Pilot Mapping"]
    headers = [normalize(mapping.cell(1, column).value) for column in range(1, 19)]
    mapping_count = 0
    hyperlink_count = 0
    for row_number in range(2, 833):
        values = [workbook_cell_value(mapping.cell(row_number, column)) for column in range(1, 19)]
        if not any(normalize(value) for value in values):
            continue
        row = {header: normalize(value) for header, value in zip(headers, values) if header}
        mapping_count += 1
        if any(mapping.cell(row_number, column).hyperlink for column in range(1, 19)):
            hyperlink_count += 1
        content = row_content(headers, values)
        evidence_id = add_row_evidence(
            "CX Pilot Mapping",
            row_number,
            row_number,
            content,
            row_fields(headers, values),
        )
        chunk_id = builder.add_chunk(
            document_id,
            f"xlsx:CXPilotMapping:r{row_number}",
            content,
            "industry_workbook",
            "mapping_guidance",
            f"CX Pilot Mapping row {row_number}",
            {
                "component": row.get("Identified Component", ""),
                "agency": row.get("Agency", ""),
                "discipline": row.get("Suggested Discipline", ""),
                "ifcEntity": row.get("IFC4 Entities", ""),
                "ifcSubtype": row.get("IFC subtypes (* USERDEFINED)", ""),
                "propertySet": row.get("Property Set", ""),
                "propertyName": row.get("Property Name", ""),
            },
            evidence_ids=[evidence_id] if evidence_id else [],
            sheet_name="CX Pilot Mapping",
            row_start=row_number,
            row_end=row_number,
        )
        add_mapping_graph(builder, chunk_id, row)

    for row_number in range(833, mapping.max_row + 1):
        values = [workbook_cell_value(mapping.cell(row_number, column)) for column in range(1, 19)]
        if any(normalize(value) for value in values):
            builder.warnings.append(
                {"source": "workbook", "sheet": "CX Pilot Mapping", "row": row_number, "message": "Row outside the declared business filter was quarantined."}
            )

    space_values = workbook["Space Values"]
    grouped_values: dict[str, list[tuple[int, str, str, str]]] = defaultdict(list)
    for row_number in range(2, space_values.max_row + 1):
        group = normalize(space_values.cell(row_number, 1).value)
        value = normalize(space_values.cell(row_number, 2).value)
        qualifier = normalize(space_values.cell(row_number, 3).value)
        if not group or not value:
            continue
        if group not in SPACE_VALUE_GROUPS:
            builder.warnings.append({"source": "workbook", "sheet": "Space Values", "row": row_number, "message": f"Unknown value group: {group}"})
        evidence_content = f"Value domain: {group}. Value: {value}{f'. Qualifier: {qualifier}' if qualifier else ''}"
        evidence_id = add_row_evidence(
            "Space Values",
            row_number,
            row_number,
            evidence_content,
            row_fields(["Value domain", "Value", "Qualifier"], [group, value, qualifier]),
        )
        grouped_values[group].append((row_number, value, qualifier, evidence_id or ""))
        domain_node = builder.add_node("ValueDomain", group, group)
        value_node = builder.add_node("AllowedValue", f"{group}:{value}:{qualifier}", value, {"qualifier": qualifier, "row": row_number})
        builder.add_edge(domain_node, value_node, "HAS_ALLOWED_VALUE", "controlled_value", None, {"qualifier": qualifier})

    for group, entries in grouped_values.items():
        for part_index in range(0, len(entries), 35):
            part = entries[part_index : part_index + 35]
            start_row, end_row = part[0][0], part[-1][0]
            content = f"Value domain: {group}. " + "; ".join(
                f"{value}{f' ({qualifier})' if qualifier else ''}" for _, value, qualifier, _ in part
            )
            chunk_id = builder.add_chunk(
                document_id,
                f"xlsx:SpaceValues:{group}:{start_row}-{end_row}",
                content,
                "industry_workbook",
                "controlled_value",
                f"Space Values rows {start_row}–{end_row}",
                {"valueDomain": group, "valueCount": len(part)},
                evidence_ids=[evidence_id for _, _, _, evidence_id in part if evidence_id],
                sheet_name="Space Values",
                row_start=start_row,
                row_end=end_row,
            )
            domain_node = builder.add_node("ValueDomain", group, group)
            builder.add_edge(document_node, domain_node, "CONTAINS", "structure", chunk_id)

    builder.documents[-1]["metadata"].update(
        {
            "mappingRowCount": mapping_count,
            "spaceValueRowCount": sum(len(values) for values in grouped_values.values()),
            "changeLogEntryCount": max(0, change_log.max_row - 1),
            "mappingHyperlinkRowCount": hyperlink_count,
            "quarantinedMappingRows": [833, 834],
            "excludedColumns": "S:W",
        }
    )
    workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", type=Path, default=DEFAULT_PDF)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--embedding-model", default="text-embedding-3-small")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    pdf_bytes = args.pdf.read_bytes()
    xlsx_bytes = args.xlsx.read_bytes()
    pdf_hash = sha256_bytes(pdf_bytes)
    xlsx_hash = sha256_bytes(xlsx_bytes)
    fingerprint = sha256_bytes(f"{pdf_hash}:{xlsx_hash}:{args.embedding_model}:extractor-v3".encode("utf-8"))
    revision_id = stable_id("corenet-x", fingerprint)
    builder = CorpusBuilder(revision_id, args.embedding_model)
    extract_pdf(builder, args.pdf, pdf_hash)
    extract_workbook(builder, args.xlsx, xlsx_hash)

    args.out.mkdir(parents=True, exist_ok=True)
    write_jsonl(args.out / "documents.jsonl", builder.documents)
    write_jsonl(args.out / "evidence.jsonl", builder.evidence)
    write_jsonl(args.out / "chunks.jsonl", builder.chunks)
    write_jsonl(args.out / "nodes.jsonl", builder.nodes.values())
    write_jsonl(args.out / "edges.jsonl", builder.edges.values())
    write_jsonl(args.out / "warnings.jsonl", builder.warnings)
    manifest = {
        "schemaVersion": 1,
        "extractorVersion": "3",
        "revisionId": revision_id,
        "sourceFingerprint": fingerprint,
        "embeddingModel": args.embedding_model,
        "embeddingDimensions": 1536,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sources": {"pdf": {"path": str(args.pdf), "sha256": pdf_hash}, "xlsx": {"path": str(args.xlsx), "sha256": xlsx_hash}},
        "counts": {
            "documents": len(builder.documents),
            "evidence": len(builder.evidence),
            "chunks": len(builder.chunks),
            "nodes": len(builder.nodes),
            "edges": len(builder.edges),
            "warnings": len(builder.warnings),
        },
    }
    (args.out / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()
